import os, json, base64, shutil, httpx, psycopg2, psycopg2.extras
from datetime import datetime, date
from dateutil import parser as dateparser
from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import FileResponse, HTMLResponse
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
import io, re

app = FastAPI(title="Sistema de Asistencia Escolar")

OLLAMA_URL   = os.getenv("OLLAMA_URL", "http://vllm-vision:8000")
VISION_MODEL = os.getenv("VISION_MODEL", "Qwen/Qwen2.5-VL-7B-Instruct")
OUTPUT_DIR   = "/app/output"
PLANTILLA    = "/app/plantilla_asistencia.xlsx"
DB_URL       = os.getenv("DATABASE_URL", "postgresql://agente:local_only_password@postgres:5432/agentes")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── DB ──────────────────────────────────────────────────────────────────────

def get_db():
    return psycopg2.connect(DB_URL)

# ─── MAPA DE LA PLANTILLA ────────────────────────────────────────────────────

MES_NUMERO = {
    "ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
    "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12
}

def get_column_map(ws):
    col_map = {}
    mes_actual = None
    for c in range(1, ws.max_column + 1):
        v7  = ws.cell(row=7,  column=c).value
        v10 = ws.cell(row=10, column=c).value
        if v7 and isinstance(v7, str) and v7.upper() in MES_NUMERO:
            mes_actual = v7.upper()
            col_map[mes_actual] = {}
        if mes_actual and v10 and isinstance(v10, (int, float)):
            dia = int(v10)
            if dia not in col_map[mes_actual]:
                col_map[mes_actual][dia] = c
    return col_map

def get_student_row_map(ws):
    row_map = {}
    for r in range(11, 41):
        nombre = ws.cell(row=r, column=2).value
        if nombre and isinstance(nombre, str):
            key = normalize_name(nombre)
            row_map[key] = r
    return row_map

def normalize_name(name: str) -> str:
    return re.sub(r'\s+', ' ', name.strip().upper())

def match_student(ocr_name: str, row_map: dict) -> str | None:
    ocr_norm = normalize_name(ocr_name)
    if ocr_norm in row_map:
        return ocr_norm
    ocr_words = set(ocr_norm.split())
    best, best_score = None, 0
    for key in row_map:
        key_words = set(key.split())
        common = len(ocr_words & key_words)
        if common >= 2 and common > best_score:
            best, best_score = key, common
    return best

# ─── OCR ─────────────────────────────────────────────────────────────────────

PROMPT_OCR = """Analiza esta foto de un registro de asistencia escolar escrito a mano.
Extrae la fecha si aparece, y todos los nombres con su tipo de inasistencia.

Responde ÚNICAMENTE con JSON válido sin markdown:
{
  "fecha": "DD/MM/YYYY o sin_fecha",
  "registros": [
    {"nombre": "APELLIDO NOMBRE", "tipo": "A"},
    {"nombre": "APELLIDO NOMBRE", "tipo": "AT"},
    {"nombre": "APELLIDO NOMBRE", "tipo": "J"}
  ]
}

Tipos válidos: A (ausente), AT (atraso), J (justificado).
Si no se especifica el tipo, asume A (ausente)."""

def imagen_a_base64(imagen_bytes: bytes) -> str:
    img = Image.open(io.BytesIO(imagen_bytes))
    if img.mode != "RGB":
        img = img.convert("RGB")
    max_size = 1024
    if max(img.size) > max_size:
        img.thumbnail((max_size, max_size), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return base64.b64encode(buf.getvalue()).decode("utf-8")

async def llamar_vision(imagen_b64: str) -> dict:
    """Llama a vLLM con la API OpenAI-compatible"""
    payload = {
        "model": VISION_MODEL,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{imagen_b64}"
                        }
                    },
                    {
                        "type": "text",
                        "text": PROMPT_OCR
                    }
                ]
            }
        ],
        "max_tokens": 1024,
        "temperature": 0.1
    }
    async with httpx.AsyncClient(timeout=1800.0) as client:
        resp = await client.post(f"{OLLAMA_URL}/v1/chat/completions", json=payload)
        if resp.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Error vLLM: {resp.text}")

    texto = resp.json()["choices"][0]["message"]["content"].strip()
    if texto.startswith("```"):
        texto = "\n".join(texto.split("\n")[1:-1])
    try:
        return json.loads(texto)
    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail=f"Modelo no devolvió JSON válido: {texto[:300]}")

# ─── EXPORTACIÓN EXCEL ───────────────────────────────────────────────────────

FILL_A  = PatternFill("solid", fgColor="FFC7CE")
FILL_AT = PatternFill("solid", fgColor="FFEB9C")
FILL_J  = PatternFill("solid", fgColor="C6EFCE")

def exportar_excel(curso_id: int, fecha_desde: date, fecha_hasta: date) -> str:
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute("SELECT * FROM cursos WHERE id = %s", (curso_id,))
    curso = cur.fetchone()
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")

    cur.execute("""
        SELECT e.numero, e.nombre, a.fecha, a.tipo
        FROM asistencia a
        JOIN estudiantes e ON e.id = a.estudiante_id
        WHERE e.curso_id = %s AND a.fecha BETWEEN %s AND %s
        ORDER BY e.numero, a.fecha
    """, (curso_id, fecha_desde, fecha_hasta))
    registros = cur.fetchall()
    cur.close()
    conn.close()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"asistencia_{curso_id}_{ts}.xlsx")
    shutil.copy2(PLANTILLA, output_path)

    wb = load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_map  = get_column_map(ws)
        row_map  = get_student_row_map(ws)

        for reg in registros:
            fecha_reg: date = reg['fecha']
            mes_nombre = list(MES_NUMERO.keys())[fecha_reg.month - 1]
            dia = fecha_reg.day

            if mes_nombre not in col_map:
                continue
            if dia not in col_map[mes_nombre]:
                continue
            col = col_map[mes_nombre][dia]

            nombre_norm = normalize_name(reg['nombre'])
            matched = match_student(nombre_norm, row_map)
            if not matched:
                continue
            row = row_map[matched]

            cell = ws.cell(row=row, column=col)
            cell.value = reg['tipo']
            cell.alignment = Alignment(horizontal='center')
            if reg['tipo'] == 'A':
                cell.fill = FILL_A
            elif reg['tipo'] == 'AT':
                cell.fill = FILL_AT
            elif reg['tipo'] == 'J':
                cell.fill = FILL_J

    wb.save(output_path)
    return output_path

# ─── ENDPOINTS ───────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index():
    with open("/app/static/index.html") as f:
        return f.read()

@app.get("/health")
async def health():
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            r = await client.get(f"{OLLAMA_URL}/health")
            vllm_ok = r.status_code == 200
    except:
        vllm_ok = False
    return {"status": "ok" if vllm_ok else "degraded", "vllm": vllm_ok, "modelo": VISION_MODEL}

@app.get("/cursos")
def listar_cursos():
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT * FROM cursos ORDER BY nombre")
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return rows

@app.post("/cursos")
def crear_curso(nombre: str = Form(...), jornada: str = Form("MATUTINA"), docente: str = Form("")):
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO cursos (nombre, jornada, docente) VALUES (%s,%s,%s) RETURNING id",
                (nombre.upper(), jornada.upper(), docente))
    id_ = cur.fetchone()[0]; conn.commit(); cur.close(); conn.close()
    return {"id": id_, "nombre": nombre}

@app.get("/cursos/{curso_id}/estudiantes")
def listar_estudiantes(curso_id: int):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT * FROM estudiantes WHERE curso_id=%s AND activo=TRUE ORDER BY numero", (curso_id,))
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return rows

@app.post("/cursos/{curso_id}/estudiantes")
def agregar_estudiante(curso_id: int, numero: int = Form(...), nombre: str = Form(...)):
    conn = get_db(); cur = conn.cursor()
    cur.execute("""INSERT INTO estudiantes (numero, nombre, curso_id) VALUES (%s,%s,%s)
                   ON CONFLICT (numero, curso_id) DO UPDATE SET nombre=EXCLUDED.nombre
                   RETURNING id""",
                (numero, nombre.upper().strip(), curso_id))
    id_ = cur.fetchone()[0]; conn.commit(); cur.close(); conn.close()
    return {"id": id_, "numero": numero, "nombre": nombre}

@app.delete("/estudiantes/{estudiante_id}")
def eliminar_estudiante(estudiante_id: int):
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE estudiantes SET activo=FALSE WHERE id=%s", (estudiante_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@app.get("/cursos/{curso_id}/asistencia")
def ver_asistencia(curso_id: int, desde: str = None, hasta: str = None):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    q = """SELECT e.numero, e.nombre, a.fecha, a.tipo, a.observacion, a.id as asistencia_id
           FROM asistencia a JOIN estudiantes e ON e.id=a.estudiante_id
           WHERE e.curso_id=%s"""
    params = [curso_id]
    if desde:
        q += " AND a.fecha >= %s"; params.append(desde)
    if hasta:
        q += " AND a.fecha <= %s"; params.append(hasta)
    q += " ORDER BY a.fecha, e.numero"
    cur.execute(q, params)
    rows = [dict(r) for r in cur.fetchall()]
    for r in rows:
        r['fecha'] = r['fecha'].isoformat()
    cur.close(); conn.close()
    return rows

@app.post("/asistencia")
def registrar_manual(
    estudiante_id: int = Form(...),
    fecha: str = Form(...),
    tipo: str = Form(...),
    observacion: str = Form("")
):
    if tipo not in ('A', 'AT', 'J'):
        raise HTTPException(400, "Tipo debe ser A, AT o J")
    conn = get_db(); cur = conn.cursor()
    cur.execute("""INSERT INTO asistencia (estudiante_id, fecha, tipo, observacion)
                   VALUES (%s,%s,%s,%s)
                   ON CONFLICT (estudiante_id, fecha)
                   DO UPDATE SET tipo=EXCLUDED.tipo, observacion=EXCLUDED.observacion, updated_at=NOW()
                   RETURNING id""",
                (estudiante_id, fecha, tipo, observacion))
    id_ = cur.fetchone()[0]; conn.commit(); cur.close(); conn.close()
    return {"id": id_}

@app.delete("/asistencia/{asistencia_id}")
def eliminar_asistencia(asistencia_id: int):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM asistencia WHERE id=%s", (asistencia_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@app.post("/procesar")
async def procesar_foto(
    foto: UploadFile = File(...),
    curso_id: int = Form(...),
    fecha_manual: str = Form("")
):
    if not foto.content_type.startswith("image/"):
        raise HTTPException(400, "Solo imágenes")
    imagen_bytes = await foto.read()
    if len(imagen_bytes) > 10 * 1024 * 1024:
        raise HTTPException(400, "Imagen muy grande (máx 10MB)")

    imagen_b64 = imagen_a_base64(imagen_bytes)
    datos = await llamar_vision(imagen_b64)

    fecha_str = fecha_manual.strip() if fecha_manual.strip() else datos.get("fecha", "")
    try:
        if fecha_str and fecha_str != "sin_fecha":
            fecha_uso = dateparser.parse(fecha_str, dayfirst=True).date()
        else:
            fecha_uso = date.today()
    except:
        fecha_uso = date.today()

    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute("SELECT id, nombre FROM estudiantes WHERE curso_id=%s AND activo=TRUE", (curso_id,))
    estudiantes = {normalize_name(r['nombre']): r['id'] for r in cur.fetchall()}

    creados, no_encontrados = 0, []
    for reg in datos.get("registros", []):
        nombre = reg.get("nombre", "")
        tipo   = reg.get("tipo", "A").upper()
        if tipo not in ("A", "AT", "J"):
            tipo = "A"

        matched = match_student(nombre, estudiantes)
        if matched:
            est_id = estudiantes[matched]
            cur.execute("""INSERT INTO asistencia (estudiante_id, fecha, tipo, foto_origen)
                           VALUES (%s,%s,%s,%s)
                           ON CONFLICT (estudiante_id, fecha)
                           DO UPDATE SET tipo=EXCLUDED.tipo, updated_at=NOW()""",
                        (est_id, fecha_uso, tipo, foto.filename))
            creados += 1
        else:
            no_encontrados.append(nombre)

    cur.execute("""INSERT INTO fotos_procesadas (filename, fecha_lista, curso_id, registros_creados, registros_no_encontrados)
                   VALUES (%s,%s,%s,%s,%s)""",
                (foto.filename, fecha_uso, curso_id, creados, no_encontrados))
    conn.commit(); cur.close(); conn.close()

    return {
        "fecha": fecha_uso.isoformat(),
        "registros_creados": creados,
        "no_encontrados": no_encontrados,
        "total_en_foto": len(datos.get("registros", []))
    }

@app.get("/exportar/excel")
def exportar(curso_id: int, desde: str, hasta: str):
    try:
        f_desde = date.fromisoformat(desde)
        f_hasta = date.fromisoformat(hasta)
    except:
        raise HTTPException(400, "Formato de fecha inválido. Usar YYYY-MM-DD")

    path = exportar_excel(curso_id, f_desde, f_hasta)
    filename = os.path.basename(path)
    return FileResponse(
        path=path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
