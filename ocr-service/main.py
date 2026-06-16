import os, json, base64, shutil, httpx, psycopg2, psycopg2.extras
from datetime import datetime
from datetime import date as date_class
from dateutil import parser as dateparser
from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import FileResponse
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
import io, re

app = FastAPI(title="OCR Service — School Attendance System")

VLLM_URL     = os.getenv("VLLM_URL", "http://vllm-vision:8000")
VISION_MODEL = os.getenv("VISION_MODEL", "Qwen/Qwen2.5-VL-7B-Instruct")
DB_URL       = os.getenv("DATABASE_URL", "postgresql://attendance:asistencia_local_2026@postgres:5432/attendance")
OUTPUT_DIR   = "/app/output"
PLANTILLA    = "/app/plantilla_asistencia.xlsx"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─── DB ──────────────────────────────────────────────────────────────────────

def get_db():
    return psycopg2.connect(DB_URL)

# ─── NORMALIZACIÓN Y MATCHING ─────────────────────────────────────────────────

def normalize_name(name: str) -> str:
    return re.sub(r'\s+', ' ', name.strip().upper())

def match_name(ocr_name: str, candidates: dict) -> str | None:
    ocr_norm = normalize_name(ocr_name)
    if ocr_norm in candidates:
        return ocr_norm
    ocr_words = set(ocr_norm.split())
    best, best_score = None, 0
    for key in candidates:
        common = len(ocr_words & set(key.split()))
        if common >= 2 and common > best_score:
            best, best_score = key, common
    return best

# ─── MAPA DE LA PLANTILLA EXCEL ──────────────────────────────────────────────

MES_NUMERO = {
    "ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
    "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12
}

def get_column_map(ws):
    col_map = {}
    mes_actual = None
    for c in range(1, ws.max_column + 1):
        v7  = ws.cell(row=7,  column=c).value
        v10 = ws.cell(row=10, column=c).value
        if v7 and isinstance(v7, str) and v7.strip().upper() in MES_NUMERO:
            mes_actual = v7.strip().upper()
            col_map[mes_actual] = {}
        if mes_actual and v10 and isinstance(v10, (int, float)):
            dia = int(v10)
            if dia not in col_map[mes_actual]:
                col_map[mes_actual][dia] = c
    return col_map

def get_student_row_map(ws):
    row_map = {}
    for r in range(11, 41):
        nombre = ws.cell(row=r, column=2).value
        if nombre and isinstance(nombre, str):
            row_map[normalize_name(nombre)] = r
    return row_map

# ─── OCR — LLAMADA A vLLM ────────────────────────────────────────────────────

PROMPT_OCR = """Analiza esta foto de un listado de inasistencias escolar escrito a mano.
Identifica todos los nombres y su tipo:
A = ausente (falta), AT = atraso (llegó tarde).
Si no se especifica el tipo asume A.
Responde ÚNICAMENTE con JSON válido sin markdown:
{
  "date": "DD/MM/YYYY o sin_fecha",
  "records": [
    {"name": "APELLIDO NOMBRE", "type": "A"}
  ]
}"""

def imagen_a_base64(imagen_bytes: bytes) -> str:
    img = Image.open(io.BytesIO(imagen_bytes))
    if img.mode != "RGB":
        img = img.convert("RGB")
    if max(img.size) > 1024:
        img.thumbnail((1024, 1024), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return base64.b64encode(buf.getvalue()).decode("utf-8")

async def call_vllm(imagen_b64: str) -> dict:
    payload = {
        "model": VISION_MODEL,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{imagen_b64}"}},
                {"type": "text", "text": PROMPT_OCR}
            ]
        }],
        "max_tokens": 1024,
        "temperature": 0.1
    }
    async with httpx.AsyncClient(timeout=1800.0) as client:
        resp = await client.post(f"{VLLM_URL}/v1/chat/completions", json=payload)
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Error vLLM: {resp.text[:300]}")
    texto = resp.json()["choices"][0]["message"]["content"].strip()
    if texto.startswith("```"):
        texto = "\n".join(texto.split("\n")[1:-1])
    try:
        return json.loads(texto)
    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail=f"Invalid JSON from model: {texto[:300]}")

# ─── ENDPOINTS ───────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            r = await client.get(f"{VLLM_URL}/health")
            vllm_ok = r.status_code == 200
    except Exception:
        vllm_ok = False
    return {"status": "ok" if vllm_ok else "degraded", "vllm": vllm_ok, "model": VISION_MODEL}

@app.post("/process-photo")
async def process_photo(
    foto: UploadFile = File(...),
    course_id: int = Form(...),
    academic_year_id: int = Form(...),
    list_date: str = Form("", alias="date")
):
    if not foto.content_type or not foto.content_type.startswith("image/"):
        raise HTTPException(400, "Only image files are accepted")

    imagen_bytes = await foto.read()
    if len(imagen_bytes) > 15 * 1024 * 1024:
        raise HTTPException(400, "Image too large (max 15MB)")

    imagen_b64 = imagen_a_base64(imagen_bytes)
    datos = await call_vllm(imagen_b64)

    fecha_str = list_date.strip() if list_date.strip() else datos.get("date", "")
    try:
        if fecha_str and fecha_str not in ("sin_fecha", ""):
            fecha_uso = dateparser.parse(fecha_str, dayfirst=True).date()
        else:
            fecha_uso = date_class.today()
    except Exception:
        fecha_uso = date_class.today()

    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute("""
        SELECT m.id AS enrollment_id, e.name
        FROM enrollments m
        JOIN students e ON e.id = m.student_id
        WHERE m.course_id = %s
          AND m.academic_year_id = %s
          AND m.deleted_at IS NULL
          AND e.deleted_at IS NULL
    """, (course_id, academic_year_id))

    enrollments = {normalize_name(r['name']): r['enrollment_id'] for r in cur.fetchall()}

    creados, no_encontrados = 0, []
    records = datos.get("records", [])

    for reg in records:
        nombre = reg.get("name", "")
        tipo   = reg.get("type", "A").upper()
        if tipo not in ("A", "AT"):
            tipo = "A"

        matched = match_name(nombre, enrollments)
        if matched:
            enroll_id = enrollments[matched]
            cur.execute("""
                INSERT INTO absences (enrollment_id, date, type, photo_source)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (enrollment_id, date)
                DO UPDATE SET type = EXCLUDED.type, updated_at = NOW()
            """, (enroll_id, fecha_uso, tipo, foto.filename))
            creados += 1
        else:
            no_encontrados.append(nombre)

    cur.execute("""
        INSERT INTO photo_logs
          (filename, list_date, course_id, academic_year_id, records_created, records_not_found)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (foto.filename, fecha_uso, course_id, academic_year_id, creados, no_encontrados))

    conn.commit()
    cur.close()
    conn.close()

    return {
        "date": fecha_uso.isoformat(),
        "records_created": creados,
        "not_found": no_encontrados,
        "total_in_photo": len(records)
    }

@app.get("/export/excel")
def export_excel(
    course_id: int,
    academic_year_id: int,
    date_from: str,
    date_to: str
):
    try:
        f_desde = date_class.fromisoformat(date_from)
        f_hasta = date_class.fromisoformat(date_to)
    except Exception:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD")

    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute("SELECT name FROM courses WHERE id = %s", (course_id,))
    course = cur.fetchone()
    if not course:
        cur.close(); conn.close()
        raise HTTPException(404, "Course not found")

    cur.execute("""
        SELECT
          e.name              AS student_name,
          m.roster_number,
          a.date,
          a.type,
          EXISTS (
            SELECT 1 FROM justification_absences ja WHERE ja.absence_id = a.id
          ) AS is_justified
        FROM absences a
        JOIN enrollments m ON m.id = a.enrollment_id
        JOIN students e    ON e.id = m.student_id
        WHERE m.course_id = %s
          AND m.academic_year_id = %s
          AND a.date BETWEEN %s AND %s
          AND a.deleted_at IS NULL
          AND m.deleted_at IS NULL
        ORDER BY m.roster_number, a.date
    """, (course_id, academic_year_id, f_desde, f_hasta))
    registros = cur.fetchall()
    cur.close()
    conn.close()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"absences_{course_id}_{ts}.xlsx")
    shutil.copy2(PLANTILLA, output_path)

    wb = load_workbook(output_path)
    fill_a  = PatternFill("solid", fgColor="FFC7CE")
    fill_at = PatternFill("solid", fgColor="FFEB9C")
    fill_j  = PatternFill("solid", fgColor="C6EFCE")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_map = get_column_map(ws)
        row_map = get_student_row_map(ws)

        for reg in registros:
            fecha_reg = reg['date']
            mes_nombre = list(MES_NUMERO.keys())[fecha_reg.month - 1]
            dia = fecha_reg.day

            if mes_nombre not in col_map or dia not in col_map[mes_nombre]:
                continue
            col = col_map[mes_nombre][dia]

            matched = match_name(reg['student_name'], row_map)
            if not matched:
                continue
            row = row_map[matched]

            display_type = "J" if reg['is_justified'] else reg['type']
            cell = ws.cell(row=row, column=col)
            cell.value = display_type
            cell.alignment = Alignment(horizontal='center')
            if display_type == "A":
                cell.fill = fill_a
            elif display_type == "AT":
                cell.fill = fill_at
            elif display_type == "J":
                cell.fill = fill_j

    wb.save(output_path)

    safe_name = re.sub(r'[^A-Za-z0-9_]', '_', course['name'])
    return FileResponse(
        path=output_path,
        filename=f"absences_{safe_name}_{date_from}_{date_to}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
